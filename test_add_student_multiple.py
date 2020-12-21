import requests
import json
import jsonpath
import openpyxl



def test_add_student_data():
    API_URL= "http://thetestingworldapi.com/api/studentsDetails"
    f= open("B:/python/pytest01/TestCases/Test.jason",'r')
    vk = openpyxl.load_workbook("B:/python/pytest01/TestCases/testing.xlsx",'r')
    sh=vk['sheet1']
    rows =sh.max_row
    for i in range (2,rows+1):
        cell_first_name=sh.cell(rov=i,column=1)
        cell_mid_name = sh.cell(rov=i, column=2)
        cell_last_name = sh.cell(rov=i, column=3)
        cell_dob = sh.cell(rov=i, column=4)
        cell_first_name.value
        cell_mid_name.value
        cell_last_name.value
        cell_dob.value






    json_requests =json.loads((f.read()))
    response= requests.post(API_URL,json_requests)
    print(response.text)
    assert response.status_code ==201
